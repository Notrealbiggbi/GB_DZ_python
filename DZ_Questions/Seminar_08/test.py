import pandas as pd
from openpyxl import load_workbook

def class_marks():
    excel_data = pd.read_excel('6B.xlsx')
    data = pd.DataFrame(excel_data, columns=['id', 'Русский язык', "Английский",
                                             "География", 'Математика', 'Астрономия'])
    print("Итоговые оценки 6B:\n", data)


def class_marks2():
    excel_data2 = pd.read_excel('5A.xlsx')
    data2 = pd.DataFrame(excel_data2, columns=['id', 'Русский язык', "Английский",
                                               "География", 'Математика', 'Астрономия'])
    print("Итоговые оценки 5A:\n", data2)

# excel_data2 = pd.read_excel('5A.xlsx')
# data2 = pd.DataFrame(excel_data2, columns=['id', 'Русский язык', "Английский",
#                                            "География", 'Математика', 'Астрономия'])
# print("Итоговые оценки 5A:\n", data2)
#
# wb = load_workbook('5A.xlsx')
# sheet = wb.get_sheet_names('Sheet1')
#
# for i in range(1, 2):
#      print(i, sheet.cell(row=i, column=2).value)
